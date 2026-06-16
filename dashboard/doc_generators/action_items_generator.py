"""Action Items Tracker (.xlsx) — detailed, point-wise. Grounded in the transcript when available."""
import os
from datetime import date

from ._common import safe_filename
from ..expanders import action_items_content

_NAVY = "FF1F3A5F"
_WHITE = "FFFFFFFF"
_PRIORITY_FILL = {"High": "FFF8D7DA", "Medium": "FFFFF3CD", "Low": "FFD4EDDA"}


def build_action_items(analysis: dict, company: dict, output_dir: str,
                       transcript: str | None = None) -> str:
    os.makedirs(output_dir, exist_ok=True)
    client = analysis.get("client_name") or "Client"
    fname = f"ActionItems_{safe_filename(client)}_{date.today().isoformat()}.xlsx"
    out = os.path.join(output_dir, fname)

    items = action_items_content(analysis, transcript)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Action Items"
    ws.sheet_view.showGridLines = False

    cols = [("#", 5), ("Action Item", 46), ("Details", 52), ("Owner", 20),
            ("Priority", 12), ("Due Date", 16), ("Status", 12)]
    for i, (_, w) in enumerate(cols, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    thin = Side(style="thin", color="FF808080")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy = PatternFill("solid", fgColor=_NAVY)
    top = Alignment(vertical="top", wrap_text=True)

    # Title
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"Action Items — {client}  ({date.today().strftime('%d %b %Y')})"
    c.font = Font(bold=True, size=14, color=_NAVY)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24

    # Header
    hdr = 2
    for i, (name, _) in enumerate(cols, start=1):
        cell = ws.cell(row=hdr, column=i, value=name)
        cell.font = Font(bold=True, color=_WHITE)
        cell.fill = navy
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = box

    r = hdr + 1
    if not items:
        ws.merge_cells(f"A{r}:G{r}")
        ws.cell(row=r, column=1, value="No action items were identified in this meeting.").font = Font(italic=True)
    for n, it in enumerate(items, start=1):
        ws.cell(row=r, column=1, value=n).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=r, column=2, value=it.get("description", "")).alignment = top
        details = "\n".join(f"• {d}" for d in it.get("details", [])) or "—"
        ws.cell(row=r, column=3, value=details).alignment = top
        ws.cell(row=r, column=4, value=it.get("owner") or "—").alignment = top
        pr = it.get("priority", "Medium")
        pcell = ws.cell(row=r, column=5, value=pr)
        pcell.alignment = Alignment(horizontal="center", vertical="top")
        if pr in _PRIORITY_FILL:
            pcell.fill = PatternFill("solid", fgColor=_PRIORITY_FILL[pr])
        ws.cell(row=r, column=6, value=it.get("due_date") or "—").alignment = top
        scell = ws.cell(row=r, column=7, value="Open")
        scell.alignment = Alignment(horizontal="center", vertical="top")
        # estimate row height from details
        lines = 1 + len(it.get("details", []))
        ws.row_dimensions[r].height = max(20, lines * 15)
        for ci in range(1, 8):
            ws.cell(row=r, column=ci).border = box
        r += 1

    ws.freeze_panes = "A3"
    wb.save(out)
    return out
