"""
Builds an Excel (.xlsx) proposal that reproduces the Helios WorkDrive proposal
template (the "Pre-Sales Evaluation" sheet): branded header with logo + Authorized
Partner badge, navy section bands (Document Information / Application & Licenses /
Data Loading / Training Details), olive BUSINESS DESCRIPTION and Milestone Name
bands, a yellow Project Timeline row, light-blue milestone title bands each
followed by ● bulleted scope tasks, and the EVALUATER DETAILS table.

The layout is rebuilt programmatically (rather than load-and-save of a sample
file) for two reasons: openpyxl drops embedded drawings on save, and the number
of milestones varies per proposal. Exact colors, fonts, merges, column widths and
the logo/badge images were extracted from the real WorkDrive sheets.

Same content schema as the .docx builder (project_title, objective, milestones,
total_effort, total_cost) so the two output formats are interchangeable.
"""
import math
import os
import re
from datetime import date

# ---- palette (extracted from the real WorkDrive template) ----
NAVY = "FF254061"        # section band background
OLIVE = "FFC3D69B"       # project-title / BUSINESS DESCRIPTION / Milestone Name bands
LIGHTBLUE = "FF93CDDD"   # pre-sales line
YELLOW = "FFFFFF00"      # project timeline / customer category A
DARK = "FF0F243F"        # title font + EVALUATER DETAILS band
BROWN = "FF984807"       # timeline / internal-evaluation font
MS_TITLE = "FFDDEBF7"    # milestone title band (accent1 lighter 80%)
WHITE = "FFFFFFFF"
BLACK = "FF000000"

_ASSETS = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets")
_LOGO = os.path.join(_ASSETS, "helios_logo.png")
_BADGE = os.path.join(_ASSETS, "badge_partner.png")

_PROVIDER_LABELS = {"openai": "OpenAI", "gemini": "Gemini", "claude": "Claude"}


def _safe_filename(title: str) -> str:
    base = re.sub(r"[^\w\- ]", "", title or "proposal").strip().replace(" ", "_")
    return (base or "proposal")[:60]


def build_xlsx(content: dict, company: dict, output_dir: str,
               provider: str | None = None) -> str:
    os.makedirs(output_dir, exist_ok=True)
    prefix = _PROVIDER_LABELS.get((provider or "").lower(), "")
    prefix = f"{prefix}_" if prefix else ""
    filename = f"{prefix}Proposal_{_safe_filename(content.get('project_title'))}_{date.today().isoformat()}.xlsx"
    out_path = os.path.join(output_dir, filename)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.sheet_view.showGridLines = False

    # Column widths (B..H carry the content; A is a thin margin).
    widths = {"A": 2.5, "B": 56.0, "C": 24.3, "D": 22.6, "E": 17.4,
              "F": 7.7, "G": 17.6, "H": 12.7}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    thin = Side(style="thin", color="FF808080")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _fill(hexargb):
        return PatternFill("solid", fgColor=hexargb)

    def cell(coord, value=None, *, bg=None, color=BLACK, size=11, bold=False,
             italic=False, align="left", valign="center", wrap=False, border=False):
        c = ws[coord]
        if value is not None:
            c.value = value
        c.font = Font(name="Calibri", size=size, bold=bold, italic=italic, color=color)
        c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
        if bg:
            c.fill = _fill(bg)
        if border:
            c.border = box
        return c

    def merge(rng):
        ws.merge_cells(rng)

    def band(r, text, bg, *, color=WHITE, size=12, align="center", span=("B", "H")):
        merge(f"{span[0]}{r}:{span[1]}{r}")
        cell(f"{span[0]}{r}", text, bg=bg, color=color, size=size, bold=True, align=align)

    # ===================== HEADER =====================
    ws.row_dimensions[2].height = 15.8
    ws.row_dimensions[3].height = 55.5
    merge("B3:H3")
    cell("B3", (company.get("name") or "HELIOS TECH LABS").upper(),
         bg=WHITE, color=BLACK, size=36, bold=True, align="center")

    # project title band (B4:H5)
    ws.row_dimensions[4].height = 14.4
    ws.row_dimensions[5].height = 22.9
    merge("B4:H5")
    cell("B4", content.get("project_title") or "Project Proposal",
         bg=OLIVE, color=DARK, size=18, bold=True, align="center", wrap=True)

    band(6, "Pre-Sales Evaluation done basis the Information note provided by the customer",
         LIGHTBLUE, color=BLACK, size=11)

    # ----- Document Information -----
    band(7, "Document Information", NAVY)
    merge("B8:C8"); cell("B8", "Document Number", bold=True)
    merge("D8:H8"); cell("D8", "Date", bold=True)
    merge("B9:C9"); cell("B9", "", size=12)
    merge("D9:H9"); cell("D9", date.today().strftime("%m/%d/%Y"), size=12)

    # ----- Application & Licenses -----
    band(10, "Application & Licenses Details ", NAVY)
    ws.row_dimensions[11].height = 43.5
    merge("B11:C11"); cell("B11", "ZOHO Modules Required:", bold=True, size=12)
    merge("D11:H11"); cell("D11", content.get("modules") or "", size=12, wrap=True)

    # ----- Timeline -----
    timeline = content.get("total_effort") or "To be confirmed"
    band(12, f"Project Timeline:  {timeline}", YELLOW, color=BROWN, size=12)
    merge("B13:H13")
    cell("B13", "Internal Evaluation: Time Required in Followings  (Man Days)   ",
         color=BROWN, size=12, bold=True, align="center")

    # ===================== MILESTONES =====================
    r = 14
    band(r, "Milestone Name", OLIVE, color=BLACK, size=16, align="center")
    ws.row_dimensions[r].height = 28.9
    r += 1

    milestones = content.get("milestones") or []
    if not milestones:
        merge(f"B{r}:H{r}")
        cell(f"B{r}", "(No milestones provided.)", italic=True, align="center")
        r += 1
    for ms in milestones:
        if isinstance(ms, dict):
            title = ms.get("title") or "Milestone"
            scope = ms.get("scope") or []
            if not isinstance(scope, list):
                scope = [scope]
        else:
            title, scope = str(ms), []
        # title band
        merge(f"B{r}:H{r}")
        cell(f"B{r}", title, bg=MS_TITLE, color=BLACK, size=14, bold=True, align="center")
        ws.row_dimensions[r].height = 24.6
        r += 1
        # bullets (single merged cell, ● per task on its own line)
        merge(f"B{r}:H{r}")
        body = "\n".join(f"● {str(t).strip()}" for t in scope) if scope else ""
        cell(f"B{r}", body, size=12, wrap=True, valign="top")
        ws.row_dimensions[r].height = _estimate_height(scope)
        r += 1

    # ===================== EVALUATER DETAILS =====================
    band(r, "EVALUATER DETAILS ", DARK)
    r += 1
    headers = [("B", "Name of Evaluator"), ("C", "Signature (Digital)"),
               ("D", "Comments (If any):"), ("E", "Requester")]
    for col, text in headers:
        cell(f"{col}{r}", text, bold=True, size=11, border=True)
    merge(f"F{r}:H{r}")
    cell(f"F{r}", "Reviewed by: (Name)", bold=True, size=11, border=True)
    r += 1
    # name row spans 3 rows tall, matching the template
    evaluator = os.getenv("PROPOSAL_EVALUATOR_NAME", "Ankush Thakur")
    reviewer = os.getenv("PROPOSAL_REVIEWER_NAME", "Jasleen Kaur")
    for col in ("B", "C", "D", "E"):
        merge(f"{col}{r}:{col}{r + 2}")
        cell(f"{col}{r}", "", border=True, valign="center")
    cell(f"B{r}", evaluator, bold=True, size=14, border=True)
    merge(f"F{r}:H{r + 2}")
    cell(f"F{r}", reviewer, bold=True, size=14, border=True)
    for rr in (r, r + 1, r + 2):
        ws.row_dimensions[rr].height = 18.0
    last_row = r + 2

    # ===================== BORDERS =====================
    # Outline every division (bands, fields, milestone cells, evaluator table)
    # by bordering each cell in the content region. Bordering all constituent
    # cells of a merged range makes Excel draw a clean box around it.
    for rr in range(4, last_row + 1):
        for cc in range(2, 9):  # columns B..H
            ws.cell(row=rr, column=cc).border = box

    # ===================== IMAGES (logo + partner badge) =====================
    _add_images(ws)

    wb.save(out_path)
    return out_path


def _estimate_height(scope) -> float:
    """Estimate a merged-cell row height (px) that fits the wrapped bullet text.

    Merged cells don't auto-fit in Excel, so we size the row ourselves. The
    content band (B..H) is ~150 character-widths wide.
    """
    if not scope:
        return 16.0
    lines = 0
    for t in scope:
        chars = len(str(t)) + 2  # "● "
        lines += max(1, math.ceil(chars / 150))
    return max(20.0, lines * 15.6 + 6)


def _add_images(ws) -> None:
    """Embed the Helios logo (top-left) and Authorized Partner badge (top-right).

    Best-effort: if Pillow or the asset files are missing, the proposal still
    builds without branding rather than failing.
    """
    try:
        from openpyxl.drawing.image import Image as XLImage
    except Exception:
        return
    if os.path.exists(_LOGO):
        try:
            logo = XLImage(_LOGO)
            logo.height = 48
            logo.width = int(48 * 2204 / 730)  # preserve native 2204x730 aspect
            ws.add_image(logo, "B3")
        except Exception:
            pass
    if os.path.exists(_BADGE):
        try:
            badge = XLImage(_BADGE)
            badge.height = 50
            badge.width = int(50 * 540 / 137)   # preserve native 540x137 aspect
            ws.add_image(badge, "G2")
        except Exception:
            pass
